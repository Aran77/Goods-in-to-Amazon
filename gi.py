# This is Program to handle String Transformations enabling us to input basic data set and extrapolate the details into an excel sheet ready for processing into various upload files for listing products on ecommerce sites such as Amazon, Ebay and Shopify

import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime


# open our data source 
df = pd.read_excel('goodsin.xlsx')
# set filename for writing data to excel
fn = r'goodsin.xlsx'

# Class defining a Parent Shoe
class shoe:
    def __init__(self, title):
        self.title = title
        self.code = ""
        self.brand = ""
        self.age_Group = ""
        self.extract_stylecode()
        self.getBrand()
        self.getAgeGroup()

    # We must extract the StyleCode String from the title
    def extract_stylecode(self):
        title = self.title.split(" ")
        # the style code always follows the word 'Trainers' in the title string, so were are being lazy and just taking the word after Trainers and calling it the style code. Regex would be a better way
        for n in range(len(title)):
            if title[n] == "Trainers":
                self.code = title[n+1]
        
    # The brand word is awlays the first word in a string. 
    def getBrand(self):   
        brand = re.split("\s+", self.title)
        print(brand)
        self.brand = brand[0]
    
    # Extract the Age Group from the string. If the chars ' GS' occur together then its a childs shoe, else its Adults
    def getAgeGroup(self):
        title = self.title.split(" ")
        for n in range(len(title)):
            if title[n].upper() == " GS":
                self.age_Group = "Kids"
            else:
                self.age_Group = "Adults"
                # Get the gener from the title string in the same way we got the Age Group.
                if title[n].lower().find(" mens"):
                    self.gender= "Mens"
                elif title[n].lower().find(" womens"):
                    self.gender = "Womens"

# Class defining the Shoe Variation, which extends the Shoe class.
class variation(shoe):
    def __init__(self, title, barcode, price, quantity, code, gender, agegroup, brand):
        self.barcode = barcode
        self.price = price
        self.quantity = quantity
        self.title = title
        self.getSizes()
        self.code = code
        self.getVariationSKU()
        self.gender = gender
        self.agegroup = agegroup
        self.brand = brand
    
    # get the varisou sizes form the title string
    def getSizes(self):
        title = self.title.split("(")
        tmptitle = title[1].split(",")
        self.size = tmptitle[0]
        tmpsizes = self.size.split(" ")
        self.UKSize = tmpsizes[1]
        self.USASize = tmpsizes[3]
        self.EUSize = tmpsizes[5]

        # get colourway formt title string
        self.colourway = tmptitle[1][:-1]
        tmpcolourcode = re.findall(r'\d+',self.colourway)
        self.colourcode = tmpcolourcode[0]

    # build the variation sku from the varisou details we have extracted from the Title string
    def getVariationSKU(self):
        if self.UKSize == "6" and self.EUSize == "39" and self.USASize.find("6.5"):
            self.sku = self.code +"-"+ self.colourcode + "-65US"
        else:
            self.sku = self.code +"-"+ self.colourcode + "-" + self.UKSize.replace(".","")
        self.parentcode = self.code

# Class to handle writing data back to the Excel file
class UpdateSheet:
    def __init__(self, relation, details, wb):
        self.relation = relation
        self.details = details
        self.wb = wb
        # open workbook and set worksheet
        self.ws = self.wb['Sheet1']     

    # Write values to variation/child lines
    def write_child(self):
        self.ws['A' + str(self.details['row'])] = self.details['sku']
        self.ws['C' + str(self.details['row'])] = self.details['code']
        self.ws['F' + str(self.details['row'])] = self.details['gender']
        self.ws['K' + str(self.details['row'])] = self.details['brand']
        self.ws['L' + str(self.details['row'])] = self.details['uksize']
        self.ws['N' + str(self.details['row'])] = self.details['usasize']
        self.ws['M' + str(self.details['row'])] = self.details['eusize']
        self.ws['o' + str(self.details['row'])] = self.details['size']
        self.ws['p' + str(self.details['row'])] = self.details['code']
        self.ws['q' + str(self.details['row'])] = self.details['colourway']      

def create_Amazon(d,row, rel,target, wb):
    ws = wb['Sheet1']
    # same value for every line
    type = "Shoes"
    # convert the row int to string
    row = str(row)
    # For UK sheet
    if target == "UK":
        # determine Browse Nodes and Gender Synonyms
        if d['gender'].lower == "mens":
            gname = "Male"
            bn = "1769797031"
        elif d['gender'].lower == "womens":
            bn = "1769860031"
            gname = "Female"
        else:
            bn = "1769675031"
            gname = "Male"
        # Start puttin gvalues in Excel cells    
        ws['A' + row] = type
        ws['b' + row] = d['code']
        ws['c' + row] = d['brand']
        ws['F' + row] = d['title']
        ws['AI' + row] = "SizeName-ColorName"
        # For variations just this
        if rel == "V":
            ws.cell(column=4, row=int(row)).number_format ='@'
            ws['d' + row] = d['barcode']
            ws['e' + row] = "EAN"
            ws['G' + row] = bn
            ws['i' + row] = "UK Footwear Size System"
            ws['J' + row] = d['agegroup']
            ws['K' + row] = "Numeric"
            ws['L' + row] = "Medium"
            ws['M' + row] = d['uksize']
            ws['N' + row] = d['colourway']
            ws['O' + row] = d['colourname']
            ws['P' + row] = 'china'
            ws['Q' + row] = d['price']
            ws['R' + row] = d['stocklevel']
            #ws['S' + row] = pic1
            ws['T' + row] = gname
            ws['U' + row] = d['agegroup']
            #ws['V' + row] = pic2
            #ws['W' + row] = pic3
            #ws['X' + row] = pic4
            #ws['Y' + row] = pic5
            #ws['Z' + row] = pic6
            ws['AF' + row] = "Child"
            ws['AG' + row] = d['code']
            ws['AH' + row] = "Variation"
            ws['AI' + row] = "SizeName-ColorName"
            ws['AI' + row] = d['gender'][:-1]
        # For parents just this
        elif rel == "P":
            ws['AF' + row] = "Parent"
            ws['B' + row] = d['code']
    
    # US VERSION
    elif target == "US":
        if d['gender'].lower == "mens":
            gname = "Male"
            #change to USA BNs ******************
            bn = "1769797031"
        elif d['gender'].lower == "womens":
            bn = "1769860031"
            gname = "Female"
        else:
            bn = "1769675031"
            gname = "Male"
        ws['A' + row] = type
    
        ws['c' + row] = d['brand']
        ws['D' + row] = d['title']
        ws['T' + row] = gname
        ws['Y' + row] = "Size/Color"
        ws['BJ' + row] = "fabric-and-synthetic"


        # For Variations do this
        if rel == "V":
            # US uses 'big_kid' not Kids.
            if d['agegroup'] == "Kids":
                agegroup = "big_kid"
            else:
                agegroup = "adult"
            ws['B' + row] = d['sku']
            ws['e' + row] = d['barcode']
            ws['f' + row] = "EAN"
            ws['G' + row] = "fashion-sneakers"
            ws['H' + row] = "Synthetic"
            ws['I' + row] = "Retail Box"
            ws['J' + row] = "US_Footwear_Size_System"
            ws['K' + row] = agegroup # Need amend to accept BIG KID'
            ws['L' + row] = "Numeric"
            ws['M' + row] = "Medium"
            ws['N' + row] = d['usasize']
            ws['BH' + row] = d['usasize']
            ws['O' + row] = d['colourway']
            ws['P' + row] = d['colourname']
            ws['V' + row] = "Child"
            ws['Q' + row] = d['price']
            ws['R' + row] = d['stocklevel']
            ws['U' + row] = agegroup
            #ws['S' + row] = pic1
            ws['w' + row] = d['code']
            ws['x' + row] = "Variation"
            ws['Z' + row] = "Unit"
            ws['AP' + row] = d['gender'][:-1]

        # For parents just this
        elif rel == "P":
            ws['B' + row] = d['code']
            ws['V' + row] = "Parent"
        
today_format = datetime.today().strftime('%d-%m-%Y')
wb = load_workbook(fn) 

# Experimental list of channels
channels = ["UK", "US"]
channels_dict = {}
# Loop through channels and load the templates
for x in channels:
    channels_dict[f"{x}wb"] = load_workbook(f'Amazon{x}Template.xlsx')     

cnt =1
# main for loop to interate over the loaded data set
for index, row in df.iterrows():
    cnt += 1
    print("processing line:" + str(cnt-1))
    # The data set consists of Parent Lines and Variation Lines, we must handle these slighlty differently.
    if row['Is Parent'] == "yes":
        # create the parent object
        parent_item = shoe(row['Title'])
        # create a disctionary of values for simple access
        d = {"code": parent_item.code, "title": row['Title'], "brand": parent_item.brand, "agegroup": parent_item.age_Group, "row": cnt, "gender": parent_item.gender}
        # call the update sheet function
        Up = UpdateSheet("P",d, wb)
        # loop the channels list and update the appropriate sheet
        for x in channels:        
            create_Amazon(d,cnt+2, "P", x, channels_dict[f"{x}wb"])
    else:
        item = variation(row['Title'], row['Barcode'], row['Price'], row['Stock Level'], parent_item.code, parent_item.gender, parent_item.age_Group, parent_item.brand)
        d = {"sku": item.sku, "title": item.title, "colourway": item.colourway, "size": item.size, "uksize":item.UKSize, "usasize": item.USASize, "eusize": item.EUSize,"row": cnt, "size": item.size, "code": item.code, "colorcode": item.colourcode, "gender": item.gender, "brand": parent_item.brand, "barcode": row['Barcode'], "stocklevel": row['Stock Level'], "agegroup": parent_item.age_Group, "colourname": row['Colour Map'], "price": row['Price']}
        Up = UpdateSheet("V", d, wb)
        Up.write_child()
        for x in channels:
            create_Amazon(d,cnt+2, "V", x, channels_dict[f"{x}wb"])

# loop through the channels list and save the workbooks.
for x in channels:
    channels_dict[f"{x}wb"].save(f"Amazon {x} " + today_format + ".xlsx")
    print(f'Writing {x} File...')

# Save the Proecessed Goods in Sheet
wb.save(f"Goods in {today_format}.xls")  



