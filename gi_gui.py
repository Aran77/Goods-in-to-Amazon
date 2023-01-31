# This is Program to handle String Transformations enabling us to input basic data set and extrapolate the details into an excel sheet ready for processing into various upload files for listing products on ecommerce sites such as Amazon, Ebay and Shopify
from tkinter import *
import customtkinter
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime
import subprocess

def openGoodsin():
    # open our data source 
    file_path = filedialog.askopenfilename()
    df = pd.read_excel(file_path)
    fn = file_path
    import_label.configure(text = "Processing, Please Wait")
    main(df, fn)

# Class defining a Parent Shoe
class shoe:
    def __init__(self, title):
        self.title = title
        self.code = ""
        self.brand = ""
        self.age_Group = ""
        self.extract_stylecode()
        self.getBrand()
        self.getAgeGroup()

    # We must extract the StyleCode String from the title
    def extract_stylecode(self):
        title = self.title.split(" ")
        # the style code always follows the word 'Trainers' in the title string, so were are being lazy and just taking the word after Trainers and calling it the style code. Regex would be a better way
        for n in range(len(title)):
            if title[n] == "Trainers":
                self.code = title[n+1]
        
    # The brand word is awlays the first word in a string. 
    def getBrand(self):   
        brand = re.split("\s+", self.title)
        self.brand = brand[0]
    
    # Extract the Age Group from the string. If the chars ' GS' occur together then its a childs shoe, else its Adults
    def getAgeGroup(self):
        title = self.title.split(" ")
        for n in range(len(title)):
            if title[n].upper() == " GS":
                self.age_Group = "Kids"
            else:
                self.age_Group = "Adults"
                # Get the gener from the title string in the same way we got the Age Group.
                if title[n].lower().find(" mens"):
                    self.gender= "Mens"
                elif title[n].lower().find(" womens"):
                    self.gender = "Womens"

# Class defining the Shoe Variation, which extends the Shoe class.
class variation(shoe):
    def __init__(self, title, barcode, price, quantity, code, gender, agegroup, brand):
        self.barcode = barcode
        self.price = price
        self.quantity = quantity
        self.title = title
        self.getSizes()
        self.code = code
        self.getVariationSKU()
        self.gender = gender
        self.agegroup = agegroup
        self.brand = brand
    
    # get the varisou sizes form the title string
    def getSizes(self):
        title = self.title.split("(")
        tmptitle = title[1].split(",")
        self.size = tmptitle[0]
        tmpsizes = self.size.split(" ")
        self.UKSize = tmpsizes[1]
        self.USASize = tmpsizes[3]
        self.EUSize = tmpsizes[5]

        # get colourway formt title string
        self.colourway = tmptitle[1][:-1]
        tmpcolourcode = re.findall(r'\d+',self.colourway)
        self.colourcode = tmpcolourcode[0]

    # build the variation sku from the varisou details we have extracted from the Title string
    def getVariationSKU(self):
        if self.UKSize == "6" and self.EUSize == "39" and self.USASize.find("6.5") and self.brand == "Nike":
            self.sku = self.code +"-"+ self.colourcode + "-65US"
        else:
            self.sku = self.code +"-"+ self.colourcode + "-" + self.UKSize.replace(".","")
        self.parentcode = self.code

# Class to handle writing data back to the Excel file
class UpdateSheet:
    def __init__(self, relation, details, wb):
        self.relation = relation
        self.details = details
        self.wb = wb
        # open workbook and set worksheet
        self.ws = self.wb['Sheet1']     

    # Write values to variation/child lines
    def write_child(self):
        self.ws['A' + str(self.details['row'])] = self.details['sku']
        self.ws['C' + str(self.details['row'])] = self.details['code']
        self.ws['F' + str(self.details['row'])] = self.details['gender']
        self.ws['K' + str(self.details['row'])] = self.details['brand']
        self.ws['L' + str(self.details['row'])] = self.details['uksize']
        self.ws['N' + str(self.details['row'])] = self.details['usasize']
        self.ws['M' + str(self.details['row'])] = self.details['eusize']
        self.ws['o' + str(self.details['row'])] = self.details['size']
        self.ws['p' + str(self.details['row'])] = self.details['code']
        self.ws['q' + str(self.details['row'])] = self.details['colourway']      
        self.ws['r' + str(self.details['row'])] = self.details['colourid']      

def PicURLS(s,c):
    return 


def create_Amazon(d,row, rel,target, wb):
    ws = wb['Sheet1']
    # same value for every line
    type = "Shoes"
    # convert the row int to string
    row = str(row)
    #path to our image library
    imgpath = "https://img.devmysite.uk/Product_Images/" + d['StyleID'] + "-" + d['ColourID'] + "/" +d['StyleID'] +"-"+ d['Colour ID'] + "_2D_000"   
    # For UK sheet
    if target == "UK":
        # determine Browse Nodes and Gender Synonyms
        if d['gender'].lower == "mens":
            gname = "Male"
            bn = "1769797031"
        elif d['gender'].lower == "womens":
            bn = "1769860031"
            gname = "Female"
        else:
            bn = "1769675031"
            gname = "Male"
        
        # Start puttin gvalues in Excel cells    
        ws['A' + row] = type
        ws['B' + row] = d['code']
        ws['C' + row] = d['brand']
        ws['F' + row] = d['title']
        ws['AI' + row] = "SizeName-ColorName"
        # For variations just this
        if rel == "V":
            ws.cell(column=4, row=int(row)).number_format ='@'
            ws['D' + row] = d['barcode']
            ws['E' + row] = "EAN"
            ws['G' + row] = bn
            ws['I' + row] = "UK Footwear Size System"
            ws['J' + row] = d['agegroup']
            ws['K' + row] = "Numeric"
            ws['L' + row] = "Medium"
            ws['M' + row] = d['uksize']
            ws['N' + row] = d['colourway']
            ws['O' + row] = d['colourname']
            ws['P' + row] = 'china'
            ws['Q' + row] = d['price']
            ws['R' + row] = d['stocklevel']
            ws['S' + row] = imgpath + "1.jpg"
            ws['T' + row] = gname
            ws['U' + row] = d['agegroup']
            ws['V' + row] = imgpath + "2.jpg"
            ws['W' + row] = imgpath + "3.jpg"
            ws['X' + row] = imgpath + "4.jpg"
            ws['Y' + row] = imgpath + "5.jpg"
            ws['Z' + row] = imgpath + "7.jpg"
            ws['AF' + row] = "Child"
            ws['AG' + row] = d['code']
            ws['AH' + row] = "Variation"
            ws['AI' + row] = "SizeName-ColorName"
            ws['AI' + row] = d['gender'][:-1]
        # For parents just this
        elif rel == "P":
            ws['AF' + row] = "Parent"
            ws['B' + row] = d['code']
    
    # US VERSION
    elif target == "US":
        if d['gender'].lower == "mens":
            gname = "Male"
            #change to USA BNs ******************
            bn = "1769797031"
        elif d['gender'].lower == "womens":
            bn = "1769860031"
            gname = "Female"
        else:
            bn = "1769675031"
            gname = "Male"
        ws['A' + row] = type
        ws['C' + row] = d['brand']
        ws['D' + row] = d['title']
        ws['T' + row] = gname
        ws['Y' + row] = "Size/Color"
        ws['BJ' + row] = "fabric-and-synthetic"

        # For Variations do this
        if rel == "V":
            price = (d['price'] + 5) * 1.16
            # US uses 'big_kid' not Kids.
            if d['agegroup'] == "Kids":
                agegroup = "big_kid"
            else:
                agegroup = "adult"
            ws['B' + row] = d['sku']
            ws['E' + row] = str(d['barcode'])
            ws['F' + row] = "EAN"
            ws['G' + row] = "fashion-sneakers"
            ws['H' + row] = "Synthetic"
            ws['I' + row] = "Retail Box"
            ws['J' + row] = "US_Footwear_Size_System"
            ws['K' + row] = agegroup # Need amend to accept BIG KID'
            ws['L' + row] = "Numeric"
            ws['M' + row] = "Medium"
            ws['N' + row] = d['usasize']
            ws['BH' + row] = d['usasize']
            ws['O' + row] = d['colourway']
            ws['P' + row] = d['colourname']
            ws['S' + row] = imgpath + "1.jpg"
            ws['V' + row] = "Child"
            ws['Q' + row] = price
            ws['R' + row] = d['stocklevel']
            ws['U' + row] = agegroup
            #ws['S' + row] = pic1
            ws['W' + row] = d['code']
            ws['X' + row] = "Variation"
            ws['Z' + row] = "Unit"
            ws['AP' + row] = d['gender'][:-1]

        # For parents just this
        elif rel == "P":
            ws['B' + row] = d['code']
            ws['V' + row] = "Parent"


def main(df, fn):
    today_format = datetime.today().strftime('%d-%m-%Y')
    wb = load_workbook(fn) 

    # Experimental list of channels
    channels = ["UK", "US"]
    channels_dict = {}
    # Loop through channels and load the templates
    for x in channels:
        channels_dict[f"{x}wb"] = load_workbook(f'Amazon{x}Template.xlsx')     

    cnt =1
    # main for loop to interate over the loaded data set
    for index, row in df.iterrows():
        cnt += 1
        # The data set consists of Parent Lines and Variation Lines, we must handle these slighlty differently.
        if row['Is Parent'] == "yes":
            # create the parent object
            parent_item = shoe(row['Title'])
            # create a disctionary of values for simple access
            d = {"code": parent_item.code, "title": row['Title'], "brand": parent_item.brand, "agegroup": parent_item.age_Group, "row": cnt, "gender": parent_item.gender}
            # call the update sheet function
            Up = UpdateSheet("P",d, wb)
            # loop the channels list and update the appropriate sheet
            for x in channels:        
                create_Amazon(d,cnt+2, "P", x, channels_dict[f"{x}wb"])
        else:        
            item = variation(row['Title'], row['Barcode'], row['Price'], row['Stock Level'], parent_item.code, parent_item.gender, parent_item.age_Group, parent_item.brand)
            d = {"sku": item.sku, "title": item.title, "colourway": item.colourway, "size": item.size.upper(), "uksize":item.UKSize, "usasize": item.USASize, "eusize": item.EUSize,"row": cnt, "code": item.code, "colorcode": item.colourcode, "gender": item.gender, "brand": parent_item.brand, "barcode": row['Barcode'], "stocklevel": row['Stock Level'], "agegroup": parent_item.age_Group, "colourname": row['Colour Map'], "price": row['Price'], "colourid": str(item.colourcode)}
            Up = UpdateSheet("V", d, wb)
            Up.write_child()
            for x in channels:
                create_Amazon(d,cnt+2, "V", x, channels_dict[f"{x}wb"])
   
    # loop through the channels list and save the workbooks.
    for x in channels:
        channels_dict[f"{x}wb"].save(f"Amazon {x} " + today_format + ".xlsx")
        print(f'Writing {x} File...')
        import_label.configure(text="UK & UK Files Saved") 
        #subprocess.Popen(r'explorer /select,"C:\python\goods in"')

    # Save the Proecessed Goods in Sheet
    wb.save(f"Goods in {today_format}.xlsx")  
    
    df1 = pd.read_excel(f"Goods in {today_format}.xlsx", dtype=str)
    # Manually setting column widths!
    col_widths = [100,0,0,0,700,50,100,50,50,50,50,50,50,50,0,50,200,0]
    tree['columns'] = df1.columns.values.tolist()
    cnt = 0
    # parse the column list into the treeview
    for i in df1.columns.values.tolist():           
        tree.column(i, width=col_widths[cnt], stretch=NO)
        tree.heading(i, text=i)
        cnt += 1
    # put the data in the treeview
    for index, row in df1.iterrows():   
        tree.insert("", 'end', text=index, values=list(row))
    # pack the tree into the tkinter layout
    tree.pack(expand=TRUE, fill=BOTH)
    
    menu_label.pack(side=TOP, padx=5, pady=5)
    sidebar_button_6.pack(side=TOP, padx=5, pady=5)

def create_EUROAmazon(d,row, rel,target, wb):
    ws = wb['Sheet1']
    # same value for every line
    type = "Shoes"
    # convert the row int to string
    row = str(row)
    #path to our photo library
    imgpath = "https://img.devmysite.uk/Product_Images/" + d['StyleID'] + "-" + d['ColourID'] + "/" +d['StyleID'] +"-"+ d['Colour ID'] + "_2D_000"    
    
    # For UK sheet
    if target == "DE":
        price = round((float(d['Price'])+15) *1.3,2)
        if price > 149.99 and price < 200:
            price = 149.99
        # determine Browse Nodes and Gender Synonyms
        if d['Gender'].lower() == "mens":
            bn = "1760376031"
            gname = "Männlich"
        elif d['Gender'].lower() == "Womens":
            bn = "1760376031"
            gname = "Weiblich"
        elif d['Gender'].lower() == "Boys":
            bn = "1760376031"
            gname = "Männlich"
        elif d['Gender'].lower() == "Girls":
            bn = "1760316031"
            gname = "Weiblich"
        else:
            bn = "1760376031"
            gname = "Männlich"
        
        # title translations
        t = {"Running Trainers": "Laufschuhe",
             "Football Boots":"Fußballschuhe",
             "Shoes":"Schuhe",
             "Basketball Trainers":"Basketball Schuhe",
             "Womens":"Damen",
             "Mens":"Herren"
             }
        for key in t:
            title = d['Title'].replace(key,t[key])
        
        # Start puttin gvalues in Excel cells    

        ws['A' + row] = type
        ws['C' + row] = d['Brand']
        ws['F' + row] = title
             
        # For variations just this
        if rel == "V":
            ws.cell(column=4, row=int(row)).number_format ='@'
            ws['B' + row] = d['Sku']
            ws['D' + row] = d['Barcode']
            ws['E' + row] = "EAN"
            ws['I' + row] = gname
            ws['K' + row] = "EU Schuhgrößensystem"
            ws['J' + row] = "Erwachsene"
            ws['L' + row] = "Erwachsene"
            ws['M' + row] = "Numerisch"
            ws['N' + row] = "Normal"
            ws['O' + row] = d['EU_Size']
            ws['Q' + row] = d['Colour Name']
            ws['R' + row] = d['Colour Map']
            ws['DD' + row] = 'china'
            ws['S' + row] = price
            ws['T' + row] = d['Stock Level']
            ws['U' + row] = imgpath + "1.jpg"
            ws['V' + row] = imgpath + "2.jpg"
            ws['W' + row] = imgpath + "3.jpg"
            ws['X' + row] = imgpath + "4.jpg"
            ws['Y' + row] = imgpath + "5.jpg"
            ws['Z' + row] = imgpath + "6.jpg"
            ws['AD' + row] = "Child"
            ws['AE' + row] = d['Parent SKU']
            ws['AF' + row] = "Variation"
            
            #ws['AI' + row] = d['gender'][:-1]
        # For parents just this
        elif rel == "P":
            ws['AD' + row] = "Parent"
            ws['B' + row] = d['Sku']
    ws['AG' + row] = "Size/Color"
    ws['G' + row] = bn

def amazon_Euros():
    today_format = datetime.today().strftime('%d-%m-%Y')
    df = pd.read_excel(f"Goods in {today_format}.xlsx", dtype=str)
    channels = ["DE", "FR", "IT", "ES"]
    channels = ["DE"]
    channels_dict = {}
    # Loop through channels and load the templates
    for x in channels:
        channels_dict[f"{x}wb"] = load_workbook(f'Amazon{x}Template.xlsx')
    cnt =1
    # main for loop to interate over the loaded data set
    for index, row in df.iterrows():
        cnt += 1
        # The data set consists of Parent Lines and Variation Lines, we must handle these slighlty differently.
        if row['Is Parent'] == "yes":
            for x in channels:
                create_EUROAmazon(row,cnt+2, "P", x, channels_dict[f"{x}wb"])
        else:        
            for x in channels:
                create_EUROAmazon(row,cnt+2, "V", x, channels_dict[f"{x}wb"])
    for x in channels:
        channels_dict[f"{x}wb"].save(f"Amazon {x} " + today_format + ".xlsx")
        print(f'Writing {x} File...')
    import_label.configure(text=f'{x} Sheet Saved')

# create Tkinter instance
app = customtkinter.CTk()
# configure window
app.title("MT Clothing Goods In")
# app.geometry(f"{1100}x{580}")
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
width = int(screen_width * 0.9)
height = int(screen_height * 0.9)
x = int((screen_width - width) / 2)
y = int((screen_height - height) / 2)
app.geometry(f"{width}x{height}+{x}+{y}")
customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("green")  # Themes: "blue" (standard), "green", "dark-blue"

# create sidebar frame with widgets
sidebar_frame = customtkinter.CTkFrame(app, width=140, corner_radius=0)
sidebar_frame.pack(side=LEFT, fill=Y)
data_frame = customtkinter.CTkFrame(app, corner_radius=0)
data_frame.pack(side=LEFT, expand=TRUE, fill=BOTH, padx=5, pady=5)
logo_label = customtkinter.CTkLabel(sidebar_frame, text="Goods In", font=customtkinter.CTkFont(size=20, weight="bold"))
logo_label.pack(side=TOP, padx=5, pady=5)
sidebar_button_1 = customtkinter.CTkButton(sidebar_frame, text="Import Goods In File", command=openGoodsin)
sidebar_button_1.pack(side=TOP, padx=5, pady=5)
status_label = customtkinter.CTkLabel(sidebar_frame, text="Status", font=customtkinter.CTkFont(size=12, weight="bold"))
status_label.pack(side=TOP, padx=5, pady=0)
import_label = customtkinter.CTkLabel(sidebar_frame, text="Open Goods In File")
import_label.pack(side=TOP, padx=5, pady=0)
#sidebar_button_2 = customtkinter.CTkButton(sidebar_frame, text="Amazon UK & US", command=create_Amazon)
#sidebar_button_2.pack(side=TOP, padx=5, pady=5)
menu_label = customtkinter.CTkLabel(sidebar_frame, text="Generate other files", font=customtkinter.CTkFont(size=12, weight="bold"))
sidebar_button_6 = customtkinter.CTkButton(sidebar_frame, text="Amazon Euros", command=amazon_Euros)

### To add later
#sidebar_button_3 = customtkinter.CTkButton(sidebar_frame, text="Shopify")
#sidebar_button_3.pack(side=TOP, padx=5, pady=5)
#sidebar_button_4 = customtkinter.CTkButton(sidebar_frame, text="Wish")
#sidebar_button_4.pack(side=TOP, padx=5, pady=5)
#sidebar_button_5 = customtkinter.CTkButton(sidebar_frame, text="Priceminister")
#sidebar_button_5.pack(side=TOP, padx=5, pady=5)
#sidebar_button_7 = customtkinter.CTkButton(sidebar_frame, text="Cdiscount")
#sidebar_button_7.pack(side=TOP, padx=5, pady=5)
#sidebar_button_8 = customtkinter.CTkButton(sidebar_frame, text="Onbuy")
#sidebar_button_8.pack(side=TOP, padx=5, pady=5)
quit_button = customtkinter.CTkButton(sidebar_frame, command=app.destroy, text="Quit")
quit_button.pack(side=BOTTOM, padx=5, pady=5)

columns = ("SKU","Title", "Barcode","Stock Level", "Price","UK Size", "US Size", "EU Size", "Colour")
# Set up Tree View for our tablular data
tree = ttk.Treeview(data_frame, columns=columns, show='headings', selectmode='none')
tree.column('#5', minwidth="400", stretch=TRUE)
tree.column('#5', minwidth="400", stretch=TRUE)
style = ttk.Style(app) 
style.theme_use("clam") # set theam to clam

style.configure('Treeview.Heading', background="", font=('Helvetica',10))
# Tkinter Loop
app.mainloop()